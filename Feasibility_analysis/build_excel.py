"""
build_excel.py — FICC Gap Analysis Workbook Generator (YAML-driven)
Replaces create_gap_analysis.py.

Reads  : ficc_data.yaml  (same directory)
Writes : FICC_Gap_Analysis.xlsx

Sheets:
  1. 差距分析总表     — 16-module gap matrix
  2. 关键功能清单     — all functions, flat, grouped by module
  3. 资源规划         — role x phase headcount table
  4. P1优先启动计划   — P1 modules with milestones
  5. 详情_<cn>        — one detail sheet per module that has function_domains
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from load_data import load as _load_data

# ── Paths ──────────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(_HERE, "FICC_Gap_Analysis.xlsx")

# ── Load data ──────────────────────────────────────────────────────────────────
_data = _load_data()

MODULES = _data["modules"]   # list of dicts
ROLES   = _data["roles"]     # list of dicts

# ── Brand palette ──────────────────────────────────────────────────────────────
NAVY    = "0F2060"
DNAV    = "1B3275"
MBLUE   = "2B5EC7"
BBLUE   = "3B7DD8"
RED     = "E53935"
CANVAS  = "F5F7FA"
CFILL   = "D6E4F7"
WHITE   = "FFFFFF"
P1COL   = "D6E4F7"   # light blue for P1 rows
P2COL   = "EEF3FA"   # very light for P2 rows
P3COL   = "F8F8F8"   # near-white for P3 rows
HDRFILL = "0F2060"   # header row background

# ── P1 launch data (presentation data not in YAML) ────────────────────────────
# keyed by module 'no'
# tuple: (start_month, milestone, first_case, risk, mitigation, team, value)
P1_LAUNCH = {
    5:  ("M1", "上线：10支债券实时监控告警，0漏报",  "某券商债券异常订单旁路检测",  "CEP引擎低延迟稳定性",    "规则引擎轻量先行，ML模型后置", "风控平台组", "监管驱动，客户愿付费，快速复制推广"),
    6:  ("M1", "上线：200+合规规则覆盖，0漏拦",      "某券商自营债券投资合规校验",  "规则库完整性",            "与合规团队联合建规则",         "风控平台组", "与异常风控组合销售，提升ARPU"),
    2:  ("M2", "上线：银行间+交易所双市场报盘",       "替换某券商现有债券交易终端",  "CFETS接口稳定性",         "早期接入测试环境联调",         "交易系统组", "替换恒生/金证，直接竞争替代"),
    3:  ("M3", "上线：正回购全周期，质押替换",        "某券商质押式回购融资管理",   "中登/上清所接口联调",     "建立完整沙箱测试套件",         "交易系统组", "与现券交易捆绑，提升粘性"),
    1:  ("M2", "上线：实时持仓净值，T0对账零差异",   "资管公司债券持仓实时账薄",   "多资产估值覆盖率",         "分批上线（先债券，后衍生品）", "平台基础组", "平台锁定效应，长期替换衡泰IBOR"),
    7:  ("M4", "上线：Brinson归因，周报自动生成",    "资管公司固收组合绩效分析",   "Brinson模型精度验证",     "对标Bloomberg绩效分析基准",    "投资管理组", "高端客户差异化，替代SimCorp/衡泰"),
    8:  ("M5", "上线：回测引擎，3个策略实盘验证",   "自营量化部门策略回测平台",   "Tick回测数据质量",         "先与AMD行情平台深度整合",      "量化研究组", "高壁垒护城河，未来SaaS化"),
    4:  ("M3", "上线：VaR/压测，监管格式报告",       "某券商FICC组合风险日报",     "模型精度vs衡泰基准",      "首期仅VaR/ES，CVA后置",        "风控平台组", "直接替代衡泰，高溢价定价"),
}

# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=None, bold=False, size=10, name="Arial"):
    kw = dict(name=name, size=size, bold=bold)
    if hex_color:
        kw["color"] = hex_color
    return Font(**kw)

def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_style(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = _fill(HDRFILL)
    c.alignment = _align("center")
    c.border = _border("4472C4")
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value, fill=None, bold=False, align="left",
              color=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color, bold=bold)
    if fill:
        c.fill = _fill(fill)
    if border:
        c.border = _border()
    c.alignment = _align(align, wrap=True)
    return c


# ── Sheet 1: 差距分析总表 ──────────────────────────────────────────────────────

def make_s1(wb):
    ws = wb.create_sheet("差距分析总表")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:P1")
    t = ws.cell(row=1, column=1, value="FICC平台建设差距分析总表  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    headers = [
        ("编号", 5), ("系统/模块", 18), ("English Name", 20), ("业务领域", 14),
        ("优先级", 7), ("市场需求\n(1-5★)", 9), ("竞品强度\n(1-5★)", 9),
        ("华锐现状\n(L0-L5)", 9), ("目标状态\n(L0-L5)", 9), ("建设差距", 8),
        ("主要竞品", 16), ("华锐基础", 18), ("工期\n(月)", 8),
        ("团队\n(人)", 8), ("人月\n投入", 8), ("战略备注", 30),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 2, col, h, width=w)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    priority_colors = {"P1": P1COL, "P2": P2COL, "P3": P3COL}
    gap_colors = {"小": "DAEEF3", "中": P2COL, "大": "FCE4D6"}

    for i, m in enumerate(MODULES):
        r = i + 3
        no         = m["no"]
        cn         = m["cn"]
        en         = m["en"]
        domain     = m["domain"]
        pri        = m["priority"]
        mkt        = m["market_demand"]
        comp       = m["competitor_strength"]
        cur        = m["current_state"]
        tgt        = m["target_state"]
        gap        = m["gap"]
        competitor = m["competitor"]
        foundation = m["foundation"]
        dur        = m["duration_m"]
        team       = m["team_sz"]
        pm         = m["person_months"]
        note       = m["note"]

        fill     = priority_colors.get(pri, P3COL)
        gap_fill = gap_colors.get(gap, P3COL)

        for col in range(1, 17):
            ws.cell(row=r, column=col).fill = _fill(fill)

        data_cell(ws, r, 1,  no,         fill, align="center")
        data_cell(ws, r, 2,  cn,         fill, bold=(pri == "P1"))
        data_cell(ws, r, 3,  en,         fill)
        data_cell(ws, r, 4,  domain,     fill)

        pc = ws.cell(row=r, column=5, value=pri)
        pc.font = Font(name="Arial", bold=True,
                       color=WHITE if pri == "P1" else DNAV, size=10)
        pc.fill = _fill(MBLUE if pri == "P1" else (BBLUE if pri == "P2" else "9DC3E6"))
        pc.alignment = _align("center")
        pc.border = _border()

        data_cell(ws, r, 6,  "★"*mkt + "☆"*(5-mkt), fill, align="center")
        data_cell(ws, r, 7,  "★"*comp + "☆"*(5-comp), fill, align="center")
        data_cell(ws, r, 8,  cur,        fill, align="center")
        data_cell(ws, r, 9,  tgt,        fill, align="center")

        gc = ws.cell(row=r, column=10, value=gap)
        gc.font = Font(name="Arial", bold=True, size=10,
                       color=RED if gap == "大" else DNAV)
        gc.fill = _fill(gap_fill)
        gc.alignment = _align("center")
        gc.border = _border()

        data_cell(ws, r, 11, competitor, fill)
        data_cell(ws, r, 12, foundation, fill)
        data_cell(ws, r, 13, dur,        fill, align="center")
        data_cell(ws, r, 14, team,       fill, align="center")

        pmc = ws.cell(row=r, column=15, value=pm)
        pmc.font = Font(name="Arial", color="0000FF", size=10)
        pmc.fill = _fill(fill)
        pmc.alignment = _align("center")
        pmc.border = _border()

        data_cell(ws, r, 16, note, fill)
        ws.row_dimensions[r].height = 28

    tr = len(MODULES) + 3
    ws.merge_cells(f"A{tr}:L{tr}")
    tc = ws.cell(row=tr, column=1, value="合计 / Totals")
    tc.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("right")
    tc.border = _border("4472C4")

    for col, formula in [
        (13, f"=SUM(M3:M{tr-1})"),
        (14, f"=MAX(N3:N{tr-1})"),
        (15, f"=SUM(O3:O{tr-1})"),
    ]:
        c = ws.cell(row=tr, column=col, value=formula)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        c.border = _border("4472C4")

    note_c = ws.cell(row=tr, column=16,
                     value="P1: 8模块(6项快赢+2项核心)  |  P2: 6模块  |  P3: 2模块")
    note_c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    note_c.fill = _fill(DNAV)
    note_c.alignment = _align()
    note_c.border = _border("4472C4")

    ws.print_area = f"A1:P{tr}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


# ── Sheet 2: 关键功能清单 ──────────────────────────────────────────────────────

def make_s2(wb):
    ws = wb.create_sheet("关键功能清单")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="FICC各模块关键功能清单  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    headers2 = [
        ("系统/模块", 18), ("优先级", 7), ("功能编号", 9), ("功能名称", 20),
        ("功能描述", 45), ("验收标准关键词", 20), ("复杂度", 8), ("预估工期(人天)", 12),
    ]
    for col, (h, w) in enumerate(headers2, 1):
        header_style(ws, 2, col, h, width=w)
    ws.row_dimensions[2].height = 26

    priority_colors = {"P1": P1COL, "P2": P2COL, "P3": P3COL}
    accept_kw = {
        "极高": "p99延迟/误差率/精度验证/压测10k TPS",
        "高":   "功能完整/边界测试/集成冒烟/性能基线",
        "中":   "功能测试通过/正确率100%",
        "低":   "功能测试通过",
    }

    row = 3
    for m in MODULES:
        pri  = m["priority"]
        fill = priority_colors.get(pri, P3COL)
        cn   = m["cn"]
        en   = m["en"]

        # Module section header (merged across all 8 columns)
        ws.merge_cells(f"A{row}:H{row}")
        label = f"{'【P1】' if pri == 'P1' else '【' + pri + '】'} {cn}  ({en})"
        gc = ws.cell(row=row, column=1, value=label)
        gc.font = Font(name="Arial", bold=True, size=10,
                       color=WHITE if pri == "P1" else DNAV)
        gc.fill = _fill(MBLUE if pri == "P1" else DNAV if pri == "P2" else "9DC3E6")
        gc.alignment = _align()
        gc.border = _border("4472C4")
        ws.row_dimensions[row].height = 20
        row += 1

        # Collect functions: domain-grouped or flat
        all_funcs = []
        if m.get("function_domains"):
            for dom in m["function_domains"]:
                for f in dom.get("functions", []):
                    all_funcs.append(f)
        else:
            all_funcs = m.get("functions", [])

        for f in all_funcs:
            complexity = f.get("complexity", "中")
            dev_days   = f.get("days", 0)
            func_id    = f.get("id", "")
            func_name  = f.get("name", "")
            func_desc  = f.get("desc", "")
            # Use acceptance field from YAML if present; else derive from complexity
            acceptance = f.get("acceptance", accept_kw.get(complexity, ""))

            data_cell(ws, row, 1, cn,         fill)
            data_cell(ws, row, 2, pri,         fill, align="center")
            data_cell(ws, row, 3, func_id,     fill, align="center")
            data_cell(ws, row, 4, func_name,   fill, bold=True)
            data_cell(ws, row, 5, func_desc,   fill)
            data_cell(ws, row, 6, acceptance,  fill)

            cc = ws.cell(row=row, column=7, value=complexity)
            cc.font = Font(name="Arial", bold=True, size=9,
                           color=WHITE if complexity == "极高" else RED if complexity == "高" else DNAV)
            cc.fill = _fill(MBLUE if complexity == "极高" else
                            "D6E4F7" if complexity == "中" else CANVAS)
            cc.alignment = _align("center")
            cc.border = _border()

            dc = ws.cell(row=row, column=8, value=dev_days)
            dc.font = Font(name="Arial", color="0000FF", size=10)
            dc.fill = _fill(fill)
            dc.alignment = _align("center")
            dc.border = _border()

            ws.row_dimensions[row].height = 28
            row += 1

    # Totals footer
    ws.merge_cells(f"A{row}:G{row}")
    tf = ws.cell(row=row, column=1, value="总计功能数 / Total Function Points")
    tf.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tf.fill = _fill(DNAV)
    tf.alignment = _align("right")
    tf.border = _border("4472C4")

    tc = ws.cell(row=row, column=8, value=f"=SUM(H3:H{row-1})")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("center")
    tc.border = _border("4472C4")


# ── Sheet 3: 资源规划 ──────────────────────────────────────────────────────────

def make_s3(wb):
    ws = wb.create_sheet("资源规划")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="FICC平台建设资源规划（三期）  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:E2")
    ws.merge_cells("F2:H2")
    ws.merge_cells("I2:K2")
    ws.merge_cells("L2:L3")

    for cell_ref, label, width in [
        ("A2", "岗位", 22), ("B2", "职责说明", 30),
        ("C2", "Phase 1（M1-M10）\nP1模块 — 核心平台建设", None),
        ("F2", "Phase 2（M11-M20）\nP2模块 — 能力扩展", None),
        ("I2", "Phase 3（M21-M27）\nP3模块 — 战略深水区", None),
        ("L2", "总招募\n峰值人力", 12),
    ]:
        c = ws[cell_ref]
        c.value = label
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        if width:
            ws.column_dimensions[cell_ref[0]].width = width

    for start_col in [3, 6, 9]:
        for j, sh in enumerate(["招募(人)", "工期(月)", "人月"]):
            col = start_col + j
            c = ws.cell(row=3, column=col, value=sh)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
            c.fill = _fill(MBLUE)
            c.alignment = _align("center")
            c.border = _border()
            ws.column_dimensions[get_column_letter(col)].width = 9

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    p1_dur, p2_dur, p3_dur = 10, 10, 7

    for i, role in enumerate(ROLES):
        r = i + 4
        role_cn = role["cn"]
        role_en = role["en"]
        p1_hc   = role["p1_hc"]
        p2_hc   = role["p2_hc"]
        p3_hc   = role["p3_hc"]
        desc    = role["desc"]

        fill = P1COL if p1_hc >= 10 else P2COL

        data_cell(ws, r, 1, f"{role_cn}\n{role_en}", fill, bold=True)
        data_cell(ws, r, 2, desc, fill)

        for start_col, hc, dur in [(3, p1_hc, p1_dur), (6, p2_hc, p2_dur), (9, p3_hc, p3_dur)]:
            hc_c = ws.cell(row=r, column=start_col, value=hc)
            hc_c.font = Font(name="Arial", color="0000FF", size=10)
            hc_c.fill = _fill(fill)
            hc_c.alignment = _align("center")
            hc_c.border = _border()

            dur_c = ws.cell(row=r, column=start_col + 1, value=dur)
            dur_c.font = Font(name="Arial", size=10)
            dur_c.fill = _fill(fill)
            dur_c.alignment = _align("center")
            dur_c.border = _border()

            hc_letter  = get_column_letter(start_col)
            dur_letter = get_column_letter(start_col + 1)
            pm_c = ws.cell(row=r, column=start_col + 2,
                           value=f"={hc_letter}{r}*{dur_letter}{r}")
            pm_c.font = Font(name="Arial", size=10)
            pm_c.fill = _fill(fill)
            pm_c.alignment = _align("center")
            pm_c.border = _border()

        max_c = ws.cell(row=r, column=12, value=f"=MAX(C{r},F{r},I{r})")
        max_c.font = Font(name="Arial", bold=True, size=10)
        max_c.fill = _fill(fill)
        max_c.alignment = _align("center")
        max_c.border = _border()

        ws.row_dimensions[r].height = 32

    tr = len(ROLES) + 4
    ws.merge_cells(f"A{tr}:B{tr}")
    tc = ws.cell(row=tr, column=1, value="合计 Totals")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("right")
    tc.border = _border("4472C4")

    for col in [3, 6, 9, 12]:
        c = ws.cell(row=tr, column=col,
                    value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr-1})")
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        c.border = _border("4472C4")

    for col in [4, 5, 7, 8, 10, 11]:
        c = ws.cell(row=tr, column=col, value="")
        c.fill = _fill(DNAV)
        c.border = _border("4472C4")

    nr = tr + 2
    ws.merge_cells(f"A{nr}:L{nr}")
    nc = ws.cell(row=nr, column=1, value=(
        "规划假设说明：① 各阶段人力可累计复用（P2团队延续P1人员，适当扩张）  "
        "② 量化研究员稀缺，建议优先招募，外部顾问补充  "
        "③ 信创合规专家建议从监管/银行方引进  "
        "④ 峰值人力约40人（Phase 2），建议2026Q3启动Phase 1招募"
    ))
    nc.font = Font(name="Arial", size=9, italic=True, color=DNAV)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[nr].height = 40


# ── Sheet 4: P1优先启动计划 ───────────────────────────────────────────────────

def make_s4(wb):
    ws = wb.create_sheet("P1优先启动计划")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value="P1重点模块优先启动计划  —  华锐技术 FICC / 2026-2027")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    headers = [
        ("模块", 18), ("启动月份", 9), ("验收里程碑", 22), ("首个客户场景", 22),
        ("核心技术风险", 22), ("风险缓解策略", 22), ("负责团队", 14),
        ("工期(月)", 8), ("团队(人)", 8), ("快赢商业价值", 22),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 2, col, h, width=w)
    ws.row_dimensions[2].height = 26

    # Sort P1 modules by launch month then by module no
    month_order = {"M1": 1, "M2": 2, "M3": 3, "M4": 4, "M5": 5}

    def sort_key(m):
        launch = P1_LAUNCH.get(m["no"])
        if launch:
            return (month_order.get(launch[0], 99), m["no"])
        return (99, m["no"])

    p1_modules = sorted(
        [m for m in MODULES if m["priority"] == "P1"],
        key=sort_key
    )

    seq = 1
    for m in p1_modules:
        launch = P1_LAUNCH.get(m["no"])
        if not launch:
            continue

        r = seq + 2
        (start, milestone, first_case, risk, mitigation, team, value) = launch
        dur = m["duration_m"]
        hc  = m["team_sz"]
        cn  = m["cn"]

        fill = P1COL
        for col in range(1, 11):
            ws.cell(row=r, column=col).fill = _fill(fill)

        data_cell(ws, r, 1, f"{seq}. {cn}", fill, bold=True)

        mc = ws.cell(row=r, column=2, value=start)
        mc.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        mc.fill = _fill(MBLUE)
        mc.alignment = _align("center")
        mc.border = _border()

        data_cell(ws, r, 3, milestone,  fill)
        data_cell(ws, r, 4, first_case, fill)

        rc = ws.cell(row=r, column=5, value=risk)
        rc.font = Font(name="Arial", size=10, color=RED)
        rc.fill = _fill(fill)
        rc.alignment = _align(wrap=True)
        rc.border = _border()

        data_cell(ws, r, 6, mitigation, fill)
        data_cell(ws, r, 7, team,       fill, align="center")

        for col, val in [(8, dur), (9, hc)]:
            vc = ws.cell(row=r, column=col, value=val)
            vc.font = Font(name="Arial", color="0000FF", size=10)
            vc.fill = _fill(fill)
            vc.alignment = _align("center")
            vc.border = _border()

        data_cell(ws, r, 10, value, fill, bold=True)
        ws.row_dimensions[r].height = 44
        seq += 1

    # Strategy summary box
    ir = seq + 2 + 1
    ws.merge_cells(f"A{ir}:J{ir}")
    ic = ws.cell(row=ir, column=1, value=(
        "P1阶段战略重心：① 快赢先行（异常风控+事前合规M1并行，6个月见成果）  "
        "② 主干平行（IBOR+现券M2启动，建立账薄主干）  "
        "③ 差异化（量化策略+绩效归因M4-M5，构建高壁垒能力）  "
        "④ 替代衡泰（FICC风险计量M3启动，以信创+精度优势直攻垄断盲区）"
    ))
    ic.font = Font(name="Arial", size=9, bold=True, color=DNAV)
    ic.fill = _fill(CFILL)
    ic.alignment = Alignment(wrap_text=True, vertical="center")
    ic.border = _border(DNAV)
    ws.row_dimensions[ir].height = 52


# ── Sheet 5+: 详情_<cn> for modules with function_domains ─────────────────────

def make_detail_sheet(wb, m):
    """
    Generalized detail sheet for any module that has function_domains.
    Mirrors make_detail_portfolio() from create_gap_analysis.py but is
    driven entirely by the module dict read from ficc_data.yaml.
    """
    cn = m["cn"]
    en = m["en"]
    safe_cn = cn.replace("/", "_").replace("\\", "_")
    ws = wb.create_sheet(f"详情_{safe_cn}")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value=f"系统详情：{cn}  |  {en}")
    t.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    # Core objective row (note + optional meta fields)
    obj_parts = []
    if m.get("note"):
        obj_parts.append(f"战略定位：{m['note']}")
    if m.get("source_slide") and m["source_slide"] not in (None, "~"):
        obj_parts.append(f"来源：Slide {m['source_slide']}")
    if m.get("data_inputs"):
        obj_parts.append(f"数据输入：{' / '.join(m['data_inputs'])}")
    if m.get("foundation_layer"):
        obj_parts.append(f"量化底座：{' · '.join(m['foundation_layer'])}")
    obj_text = "   |   ".join(obj_parts) if obj_parts else f"{cn} 功能详情"

    ws.merge_cells("A2:J2")
    obj = ws.cell(row=2, column=1, value=obj_text)
    obj.font = Font(name="Arial", size=10, italic=True, color=DNAV)
    obj.fill = _fill(CFILL)
    obj.alignment = _align(wrap=True)
    obj.border = _border(DNAV)
    ws.row_dimensions[2].height = 36

    # Architecture / summary row
    arch_text = (
        f"模块：{cn}  ({en})  |  "
        f"优先级：{m['priority']}  |  "
        f"竞品：{m['competitor']}  |  "
        f"华锐基础：{m['foundation']}  |  "
        f"工期：{m['duration_m']}月  团队：{m['team_sz']}人"
    )
    ws.merge_cells("A3:J3")
    arch = ws.cell(row=3, column=1, value=arch_text)
    arch.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    arch.fill = _fill(DNAV)
    arch.alignment = _align(wrap=True)
    arch.border = _border("4472C4")
    ws.row_dimensions[3].height = 28

    # Column headers
    headers = [
        ("功能域", 16), ("编号", 8), ("功能名称", 22), ("功能描述", 48),
        ("验收标准", 26), ("集成依赖", 20), ("复杂度", 8),
        ("预估(人天)", 10), ("优先级", 7), ("备注", 18),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 4, col, h, width=w)
    ws.row_dimensions[4].height = 28

    # Domain fill palette — cycle through brand blues
    domain_fills = [MBLUE, DNAV, BBLUE, "3B7DD8", "9DC3E6", CANVAS]
    function_domains = m.get("function_domains", [])

    row = 5
    for di, dom in enumerate(function_domains):
        dom_name = f"{dom['id']}. {dom['name']}"
        dom_fill = domain_fills[di % len(domain_fills)]
        is_dark  = dom_fill not in (CANVAS, "9DC3E6")

        # Domain header row
        ws.merge_cells(f"A{row}:J{row}")
        dh = ws.cell(row=row, column=1, value=dom_name)
        dh.font = Font(name="Arial", bold=True, size=10,
                       color=WHITE if is_dark else DNAV)
        dh.fill = _fill(dom_fill)
        dh.alignment = _align()
        dh.border = _border("4472C4")
        ws.row_dimensions[row].height = 18
        row += 1

        for f in dom.get("functions", []):
            fid        = f.get("id", "")
            fname      = f.get("name", "")
            fdesc      = f.get("desc", "")
            accept     = f.get("acceptance", "")
            dep        = f.get("dependency", "")
            complexity = f.get("complexity", "中")
            days       = f.get("days", 0)
            priority   = f.get("priority", m["priority"])
            note       = f.get("note", "")

            row_fill = P1COL if priority == "P1" else P2COL

            data_cell(ws, row, 1, dom_name, row_fill)
            data_cell(ws, row, 2, fid,      row_fill, align="center")
            data_cell(ws, row, 3, fname,    row_fill, bold=True)
            data_cell(ws, row, 4, fdesc,    row_fill)
            data_cell(ws, row, 5, accept,   row_fill)
            data_cell(ws, row, 6, dep,      row_fill)

            cc = ws.cell(row=row, column=7, value=complexity)
            cc.font = Font(name="Arial", bold=True, size=9,
                           color=WHITE if complexity == "极高" else RED if complexity == "高" else DNAV)
            cc.fill = _fill(MBLUE if complexity == "极高" else P2COL)
            cc.alignment = _align("center")
            cc.border = _border()

            dc = ws.cell(row=row, column=8, value=days)
            dc.font = Font(name="Arial", color="0000FF", size=10)
            dc.fill = _fill(row_fill)
            dc.alignment = _align("center")
            dc.border = _border()

            pc = ws.cell(row=row, column=9, value=priority)
            pc.font = Font(name="Arial", bold=True, size=10,
                           color=WHITE if priority == "P1" else DNAV)
            pc.fill = _fill(MBLUE if priority == "P1" else BBLUE)
            pc.alignment = _align("center")
            pc.border = _border()

            data_cell(ws, row, 10, note, row_fill)
            ws.row_dimensions[row].height = 48
            row += 1

    # Totals
    total_funcs = sum(len(dom.get("functions", [])) for dom in function_domains)
    p1_count = sum(
        1 for dom in function_domains
        for f in dom.get("functions", [])
        if f.get("priority", m["priority"]) == "P1"
    )
    p2_count = total_funcs - p1_count

    ws.merge_cells(f"A{row}:G{row}")
    tf = ws.cell(row=row, column=1,
                 value=f"合计 {total_funcs}项功能  |  P1: {p1_count}项  |  P2: {p2_count}项")
    tf.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tf.fill = _fill(DNAV)
    tf.alignment = _align()
    tf.border = _border("4472C4")

    tc = ws.cell(row=row, column=8, value=f"=SUM(H5:H{row-1})")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("center")
    tc.border = _border("4472C4")

    for col in [9, 10]:
        c = ws.cell(row=row, column=col, value="")
        c.fill = _fill(DNAV)
        c.border = _border("4472C4")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)

    make_s1(wb)
    make_s2(wb)
    make_s3(wb)
    make_s4(wb)

    # One detail sheet per module that has function_domains
    for m in MODULES:
        if m.get("function_domains"):
            make_detail_sheet(wb, m)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Saved : {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
