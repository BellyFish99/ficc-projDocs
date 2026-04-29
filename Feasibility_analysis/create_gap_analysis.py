"""
FICC Gap Analysis Workbook Generator
Creates FICC_Gap_Analysis.xlsx with 4 sheets:
  S1 差距分析总表  — 16-module matrix with gap scores, resources, timelines
  S2 关键功能清单  — function-level breakdown per module
  S3 资源规划     — headcount plan across 3 phases
  S4 P1优先启动   — P1 fast-track detail
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = "/mnt/d/work/ficc/Feasibility_analysis/FICC_Gap_Analysis.xlsx"

# ── Brand palette ─────────────────────────────────────────────────────────────
NAVY    = "0F2060"
DNAV    = "1B3275"
MBLUE   = "2B5EC7"
BBLUE   = "3B7DD8"
RED     = "E53935"
CANVAS  = "F5F7FA"
CFILL   = "D6E4F7"
WHITE   = "FFFFFF"
P1COL   = "D6E4F7"   # light blue for P1 rows
P2COL   = "EEF3FA"   # very light for P2 rows
P3COL   = "F8F8F8"   # near-white for P3 rows
HDRFILL = "0F2060"   # header row background

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=None, bold=False, size=10, name="Arial"):
    kw = dict(name=name, size=size, bold=bold)
    if hex_color:
        kw["color"] = hex_color
    return Font(**kw)

def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def header_style(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = _fill(HDRFILL)
    c.alignment = _align("center")
    c.border = _border("4472C4")
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, value, fill=None, bold=False, align="left",
              color=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = _font(color, bold=bold)
    if fill:
        c.fill = _fill(fill)
    if border:
        c.border = _border()
    c.alignment = _align(align, wrap=True)
    return c


# ── Data ──────────────────────────────────────────────────────────────────────

MODULES = [
    # (no, cn_name, en_name, domain, priority, mkt★, comp★, cur_state, tgt_state,
    #  gap, competitor, ar_foundation, duration_m, team_sz, person_months, note)
    (1,  "IBOR投资账薄管理",      "IBOR",                    "前/中台基础设施", "P1", 5, 3, "L2", "L4", "中", "衡泰(老旧)", "AMI/AMDB账户体系",         8,  10, 70,  "整个FICC平台的账薄主干，先建先通"),
    (2,  "现券交易系统",           "Bond Trading",            "前台交易",        "P1", 5, 2, "L3", "L4", "小", "恒生/金证", "ATP债券报盘引擎",            6,  8,  42,  "银行间+交易所双市场覆盖"),
    (3,  "回购管理",               "Repo Management",         "前台交易",        "P1", 5, 2, "L2", "L4", "中", "恒生/金证", "M9回购管理模块",             5,  6,  27,  "资金融通核心，与IBOR强耦合"),
    (4,  "FICC风险计量平台",       "FICC Risk Measurement",   "中台风控",        "P1", 5, 4, "L2", "L4", "中", "衡泰(垄断)", "ARC风险引擎",               10, 10, 80,  "替代衡泰老旧系统，信创切入口"),
    (5,  "异常交易风控",           "Abnormal Trade Monitor",  "中台风控",        "P1", 5, 1, "L3", "L5", "小", "无主导者",  "ARC+AI规则引擎",             6,  8,  42,  "监管驱动，AI差异化，快速商业化"),
    (6,  "事前风控/合规",          "Pre-trade Risk",          "中台风控",        "P1", 5, 2, "L3", "L4", "小", "恒生/金证", "ARC合规模块",                5,  6,  28,  "嵌入交易前拦截，合规刚需"),
    (7,  "组合管理+绩效归因",      "Portfolio Mgmt & Perf",   "前台投资管理",    "P1", 5, 3, "L2", "L4", "中", "SimCorp/衡泰", "M9 PMS模块",              12, 12, 120, "17项功能含量化底座集成；Slide44全景展开"),
    (8,  "策略投资/量化",          "Quant Strategy",          "前台投研",        "P1", 4, 2, "L1", "L3", "中", "无主导者",  "低延迟ATP执行层",            10, 12, 96,  "因子框架+回测引擎，低延迟执行"),
    (9,  "量化定价引擎",           "Quant Pricing Engine",    "中台定价",        "P2", 5, 5, "L1", "L4", "大", "衡泰(垄断)", "ARC基础数值库",             18, 15, 225, "蒙特卡洛/有限差分，替代衡泰核心"),
    (10, "利率衍生品",             "Interest Rate Derivatives","前台交易",       "P2", 4, 4, "L1", "L3", "大", "Murex/衡泰", "无直接基础",                12, 12, 108, "IRS/OIS/CCS，需专业量化团队"),
    (11, "信用债/信用衍生品",      "Credit & CDS",            "前台交易",        "P2", 4, 4, "L1", "L3", "大", "衡泰/Bloomberg", "ARC信用风险框架",         10, 10, 80,  "CDS定价+信用利差分析"),
    (12, "数据管理/行情平台",      "Market Data Platform",    "基础数据",        "P2", 4, 3, "L3", "L4", "小", "Bloomberg/Wind", "AMD行情平台",              4,  6,  22,  "多源聚合+估值价格发布，赋能其他模块"),
    (13, "交易对手管理",           "Counterparty Management", "中台信用",        "P2", 4, 3, "L1", "L3", "中大","衡泰/专项系统", "ARC信用模块",             8,  8,  56,  "ISDA/CSA管理+CVA/DVA计算"),
    (14, "清算结算后台",           "Clearing & Settlement",   "后台运营",        "P2", 4, 4, "L2", "L3", "中", "金证/中台", "ACP清算平台",                7,  8,  50,  "DVP/FOP+中登中债登接口"),
    (15, "做市商系统",             "Market Making",           "前台做市",        "P3", 3, 5, "L0", "L2", "大", "Murex/专项", "ATP报价接口",               12, 12, 120, "双边报价引擎，需深度市场经验"),
    (16, "外汇/外汇衍生品",        "FX & FX Derivatives",     "前台交易",        "P3", 3, 5, "L0", "L2", "大", "Murex/Bloomberg", "无",                   12, 10, 100, "FX/NDF市场门槛极高，建议Phase 3"),
]

# ── Key functions per module (module_no, func_no, func_cn, func_desc, complexity, dev_days)
FUNCTIONS = [
    # ─── 1. IBOR
    (1, 1,  "跨资产持仓实时净值",     "支持债券/利率衍生品/回购全品种实时估值，T+0净值计算",                                "高",   45),
    (1, 2,  "多账户层级管理",         "账户/策略/产品/公司层级合并与分拆，含多级组合结构",                                  "中",   25),
    (1, 3,  "现金流预测",             "回购到期/票息支付/衍生品结算现金流7天预测，流动性缓冲预警",                           "高",   30),
    (1, 4,  "账薄T0/T1切换",          "日内账薄（含未交割）与结算账薄自动切换，支持对账差异分析",                            "中",   20),
    (1, 5,  "托管行对账接口",         "与中登、中债登、托管银行日终对账，差异自动推送告警",                                  "高",   25),
    (1, 6,  "估值价格接入",           "接入AMD/Bloomberg/Wind估值价格服务，支持估值覆盖",                                   "中",   15),

    # ─── 2. 现券交易
    (2, 1,  "银行间双边报价",         "CFETS/X-Bond两板块报价，支持询价（RFQ）和竞价",                                       "高",   30),
    (2, 2,  "交易所债券委托",         "上交所/深交所债券买卖下单，支持限价/市价/条件单",                                     "中",   20),
    (2, 3,  "收益率/价格计算",        "全品种YTM/价格双向转换，含含权债修正久期计算",                                        "中",   15),
    (2, 4,  "交易前风险校验",         "调用事前风控接口，单笔限额/集中度/利率风险实时拦截",                                  "中",   15),
    (2, 5,  "交易确认与指令生成",     "成交确认后自动生成DVP/FOP清算指令，推送清算结算模块",                                  "高",   20),
    (2, 6,  "做市/询价接口",          "支持双边报价模式，库存自动刷新，与做市商系统预留接口",                                 "中",   10),

    # ─── 3. 回购管理
    (3, 1,  "正/逆回购全周期",        "协议回购（GC/质押式）发起、续做、到期，双方确认闭环",                                  "高",   30),
    (3, 2,  "质押品篮子管理",         "多类质押品估值与折扣率，支持替换与追加保证金（Margin Call）",                          "高",   25),
    (3, 3,  "资金头寸监控",           "日内融资头寸实时监控，资金缺口预警，与IBOR现金流对接",                                  "中",   20),
    (3, 4,  "到期自动提醒",           "到期前T-1/T-3提醒，支持批量续作规则配置",                                             "低",   10),
    (3, 5,  "CCP/上清所接口",         "标准化质押式回购通过上清所CCP清算，接口对接",                                          "高",   20),

    # ─── 4. FICC风险计量
    (4, 1,  "VaR/CVaR/ES计算",        "历史模拟法/参数法/蒙特卡洛三模式，置信度99%/99.5%可选",                               "极高", 60),
    (4, 2,  "压力测试情景库",         "内置924行情/2015股灾/2020疫情等10+情景，支持自定义",                                   "高",   40),
    (4, 3,  "限额体系与超限预警",     "交易员/策略/产品/公司4级限额，实时穿透预警，告警推送",                                  "高",   30),
    (4, 4,  "信用风险（CCR）",        "交易对手敞口实时聚合，EEPE/PFE/CVA估算",                                              "极高", 50),
    (4, 5,  "希腊字母（Greeks）",     "Delta/Gamma/Vega/Theta/DV01日内实时计算，覆盖固收+衍生品",                            "高",   35),
    (4, 6,  "风险报告自动化",         "每日风险报告（Excel/PDF）自动生成，支持监管报送格式",                                   "中",   20),

    # ─── 5. 异常交易风控
    (5, 1,  "实时CEP监控引擎",        "基于复杂事件处理（CEP），毫秒级事件流检测，滑动窗口聚合",                              "极高", 50),
    (5, 2,  "异常价格检测",           "基于历史分位数/实时行情计算价格偏离度，动态阈值自适应",                                 "高",   25),
    (5, 3,  "AI异常识别模型",         "机器学习（XGBoost/LSTM）训练异常交易识别模型，在线推理",                               "高",   40),
    (5, 4,  "规则引擎（可配置）",     "No-Code规则配置界面，合规人员自助维护检测规则",                                        "中",   20),
    (5, 5,  "监管报送自动化",         "一键生成证监会/人行格式异常交易报告，历史追溯查询",                                    "中",   25),

    # ─── 6. 事前风控/合规
    (6, 1,  "投资策略合规规则库",     "内置1000+条监管规则（投资指引/证监会规定），支持新增",                                  "高",   30),
    (6, 2,  "实时订单拦截",           "在交易指令提交前同步校验，不满足规则则拦截+原因告知",                                   "极高", 25),
    (6, 3,  "限额多维度检查",         "单笔名义/资产类别/集中度/净敞口/利率敏感度多维并行检查",                               "高",   25),
    (6, 4,  "合规审批工作流",         "超限申请→审批→记录全流程，支持移动端审批",                                             "中",   15),
    (6, 5,  "监管白/黑名单",          "证券/交易对手监管限制名单实时同步，拦截违规交易",                                       "中",   15),

    # ─── 7. 组合管理+绩效归因  (基于0417规划文档Slide 44展开)
    # A. 组合分析
    (7,  1, "债券横截面与时序分析",   "跨品种横截面信号挖掘（利差、久期偏差），历史时序趋势/均值回归分析",                    "高",   35),
    (7,  2, "相对价值分析",           "利差/Z-spread/OAS相对价值，历史分位数定位，套利信号识别",                               "高",   30),
    (7,  3, "组合相关度分析",         "资产间收益相关矩阵计算，组合集中度风险识别，分散化系数监控",                            "中",   25),
    # B. 风险计量
    (7,  4, "情景分析",               "利率/信用/流动性三类自定义情景，组合P&L影响量化，多情景并行对比",                       "高",   35),
    (7,  5, "敏感度分析",             "久期/DV01/凸度/利差敏感度多因子并行计算，组合对冲比率建议",                             "高",   25),
    (7,  6, "极端事件分析",           "历史极端情景（924行情/2015股灾/2020疫情）组合压测，尾部损失分布",                       "高",   30),
    # C. 绩效归因
    (7,  7, "指数比较分析",           "对标中债综合/信用/政策行等指数，跟踪误差（TE）实时监控，超额Alpha分解",                  "中",   25),
    (7,  8, "绩效归因（Brinson）",    "BHB模型三效应：资产配置/个券选择/交互效应，日频计算，周/月/季报自动生成",               "高",   40),
    # D. 组合优化与再平衡
    (7,  9, "虚拟组合分析",           "组合调整前后效果模拟（净值/风险/收益），交易冲击成本预估，方案比较",                    "高",   20),
    (7, 10, "再平衡方案试算与比较",   "久期/集中度/流动性约束联合优化，多方案并行试算，最优路径推荐",                          "高",   35),
    (7, 11, "资产配置策略分析",       "宏观因子驱动大类资产配置模型（利率/信用/权益），动态权重优化建议",                      "高",   40),
    (7, 12, "固收+策略分析",          "债券底仓+衍生品增强策略P&L归因，风险分解，固收+产品净值模拟",                           "高",   30),
    (7, 13, "量化策略组合分析",       "多策略组合叠加容量约束，策略相关性管理，Sharpe/最大回撤监控",                           "高",   35),
    # E. 流动性管理
    (7, 14, "流动性分析",             "逐券换手率/市场深度评估，组合变现能力分析，流动性压力下变现损失估算",                    "中",   20),
    # F. 量化计算底座集成
    (7, 15, "定价计算集成",           "调用量化定价引擎实时估值与Greeks，支持含权债/可转债，结果缓存与推送",                    "高",   20),
    (7, 16, "利率曲线集成",           "多曲线实时消费（国债/政策行/信用），曲线因子提取（水平/斜率/曲率），敏感度映射",         "中",   15),
    (7, 17, "因子模型",               "多因子风险模型（利率/信用/流动性/行业因子），Alpha因子库，因子暴露实时计算",             "高",   40),

    # ─── 8. 策略投资/量化
    (8, 1,  "因子模型框架",           "多因子Alpha模型+风险因子，支持自定义因子上传与测试",                                   "高",   40),
    (8, 2,  "Tick级回测引擎",         "Tick数据驱动回测，真实滑点/冲击模型，批量参数优化",                                    "极高", 60),
    (8, 3,  "信号生成与调度",         "策略信号实时计算，支持Python/C++策略脚本沙箱运行",                                     "高",   40),
    (8, 4,  "低延迟执行接口",         "与ATP超低延迟执行层对接，算法拆单（TWAP/VWAP/IS）",                                    "极高", 45),
    (8, 5,  "实时策略监控",           "持仓/PnL/回撤/Sharpe实时看板，风险触发自动平仓",                                       "中",   20),

    # ─── 9. 量化定价引擎
    (9, 1,  "期限结构模型",           "Nelson-Siegel/Svensson/HJM等多模型参数拟合，实时曲线",                                 "极高", 60),
    (9, 2,  "蒙特卡洛求解器",         "GPU加速MC，QMC低差异序列，支持路径相关期权定价",                                       "极高", 70),
    (9, 3,  "有限差分（PDE）",        "Crank-Nicolson/ADI方法，含权债/可转债精确定价",                                        "极高", 60),
    (9, 4,  "实时Greeks",             "Delta/Gamma/Vega/Theta/Rho分布式实时计算",                                             "高",   40),
    (9, 5,  "定价服务微服务架构",     "REST/gRPC定价API，支持横向扩展，p99<100ms",                                            "高",   40),
    (9, 6,  "历史估值回测",           "基于历史行情重放定价误差分析，模型验证框架",                                            "高",   30),

    # ─── 10. 利率衍生品
    (10, 1, "IRS/OIS/CCS定价",        "利率互换标准与非标结构定价，折现曲线OIS",                                               "极高", 60),
    (10, 2, "久期/DV01/凸度",         "利率衍生品Greeks实时计算，组合对冲比率",                                               "高",   30),
    (10, 3, "LCH/上清所CCP接口",      "标准化IRS中央清算，保证金（IM/VM）计算",                                               "高",   35),
    (10, 4, "压力测试/情景分析",      "利率曲线平移/扭转/蝶式情景，组合P&L影响",                                              "高",   30),
    (10, 5, "利率曲线发布",           "多曲线框架（OIS/Libor过渡/中债曲线）实时发布",                                         "高",   30),

    # ─── 11. 信用债/信用衍生品
    (11, 1, "信用利差分析",           "行业/评级/期限信用利差曲线实时计算，历史分位数分析",                                    "高",   30),
    (11, 2, "CDS/CLN定价",            "信用违约互换定价，ISDA标准，生存概率曲线校准",                                         "极高", 50),
    (11, 3, "信用评级动态监控",       "对接外部评级机构（中诚信/联合/大公），评级下调预警",                                    "中",   20),
    (11, 4, "违约概率（PD）建模",     "结构化Merton模型+市场隐含PD，信用风险计量",                                             "高",   40),
    (11, 5, "集中度限额管理",         "行业/发行人/评级集中度多维限额，超限预警",                                              "中",   20),

    # ─── 12. 数据管理/行情平台
    (12, 1, "多源行情聚合",           "Bloomberg/Wind/中汇信/Refinitiv多源融合，冲突解析",                                    "高",   25),
    (12, 2, "数据质量校验",           "异常值/缺失值/跳价自动检测，修复建议与人工审核",                                        "中",   15),
    (12, 3, "历史数据仓库",           "Tick/日频历史行情存储，高效查询（ClickHouse/TimescaleDB）",                             "高",   25),
    (12, 4, "收益率曲线发布",         "国债/政策行/高等级信用曲线实时发布，Kafka广播",                                         "高",   20),
    (12, 5, "估值价格服务",           "非活跃债券估值价格（中债估值基准），API接口提供",                                        "中",   15),

    # ─── 13. 交易对手管理
    (13, 1, "对手方评级与审批",       "内部信用评级模型，新增/变更交易对手审批流",                                             "中",   20),
    (13, 2, "实时敞口聚合",           "跨产品交易对手名义/MtM敞口实时聚合，压力敞口（EPE）",                                   "高",   35),
    (13, 3, "ISDA/CSA协议管理",       "主协议文档管理，净额结算协议（Netting Set）配置",                                       "高",   25),
    (13, 4, "CVA/DVA/FVA计算",        "信用/债务/资金估值调整实时估算，敞口报告",                                              "极高", 40),
    (13, 5, "集中度与限额",           "对手方/集团合并敞口限额，超限审批与记录",                                               "中",   15),

    # ─── 14. 清算结算后台
    (14, 1, "DVP/FOP指令生成",        "成交后自动生成交割指令，批量/单笔支持",                                                 "高",   25),
    (14, 2, "中登/中债登接口",        "银行间债券（中债登）和交易所债券（中登）接口对接",                                      "极高", 40),
    (14, 3, "待交割明细管理",         "T+0/T+1待交割跟踪，提前预警可能失败交割",                                               "高",   25),
    (14, 4, "资金调拨确认",           "资金划拨指令生成，银行资金头寸确认，轧差净额",                                          "高",   25),
    (14, 5, "交割失败处理",           "失败交割自动识别，补交割流程，罚息计算",                                                "中",   20),

    # ─── 15. 做市商系统
    (15, 1, "双边报价引擎",           "多品种同时双边报价，延迟<5ms，报价宽度动态优化",                                        "极高", 70),
    (15, 2, "库存风险实时管理",       "做市库存PnL实时监控，Delta对冲自动触发",                                                "极高", 60),
    (15, 3, "报价优化算法",           "基于市场微结构理论动态调整bid-ask spread",                                              "极高", 50),
    (15, 4, "竞争力分析",             "与市场最优报价对比，中签率分析，策略调优",                                               "高",   25),
    (15, 5, "交易对手管理集成",       "做市交易对手信用实时检查，额度占用同步更新",                                             "中",   20),

    # ─── 16. 外汇/外汇衍生品
    (16, 1, "即期/远期报价",          "USD/CNY等主要货币对即期汇率+远期点数报价",                                               "高",   40),
    (16, 2, "外汇期权（FXO）",        "欧式/美式FXO定价（Garman-Kohlhagen），Greeks",                                          "极高", 60),
    (16, 3, "NDF/NDO定价",            "不可交割远期/期权定价，境外对手对接",                                                   "极高", 50),
    (16, 4, "跨货币敞口汇总",         "多币种资产折算人民币敞口，汇率风险VaR",                                                  "高",   30),
    (16, 5, "CFETS/外汇局接口",       "银行间外汇市场CFETS接口，外汇局合规报送",                                               "高",   35),
]

# Resource plan: (role_cn, role_en, p1_hc, p2_hc, p3_hc, desc)
ROLES = [
    ("FICC业务架构师",  "FICC Business Architect",  2, 2, 1, "负责各模块业务需求与架构设计，需有银行间/衍生品经验"),
    ("后端研发工程师",  "Backend Engineer",          15, 20, 10, "Java/Go微服务，低延迟计算，Kafka/Redis"),
    ("量化研究员",      "Quant Researcher",          6, 8, 4, "定价模型、风险计量、因子模型，Python/C++"),
    ("前端研发工程师",  "Frontend Engineer",         4, 6, 3, "交易终端UI，React/Vue，实时数据展示"),
    ("数据工程师",      "Data Engineer",             3, 4, 2, "行情接入、数据管道、历史数据仓库"),
    ("测试工程师",      "QA Engineer",               4, 6, 3, "自动化测试，压测，合规验证"),
    ("产品经理/BA",     "Product Manager / BA",      3, 3, 2, "需求梳理、验收标准、客户沟通"),
    ("DevOps/基础设施", "DevOps / Infra",            2, 3, 2, "信创部署、K8s、监控告警体系"),
    ("安全合规专家",    "Security & Compliance",     1, 2, 1, "合规规则库维护，监管报送，数据安全"),
]


def make_s1(wb):
    ws = wb.create_sheet("差距分析总表")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:P1")
    t = ws.cell(row=1, column=1, value="FICC平台建设差距分析总表  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    # Header row 2
    headers = [
        ("编号", 5), ("系统/模块", 18), ("English Name", 20), ("业务领域", 14),
        ("优先级", 7), ("市场需求\n(1-5★)", 9), ("竞品强度\n(1-5★)", 9),
        ("华锐现状\n(L0-L5)", 9), ("目标状态\n(L0-L5)", 9), ("建设差距", 8),
        ("主要竞品", 16), ("华锐基础", 18), ("工期\n(月)", 8),
        ("团队\n(人)", 8), ("人月\n投入", 8), ("战略备注", 30),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 2, col, h, width=w)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    priority_colors = {"P1": P1COL, "P2": P2COL, "P3": P3COL}
    gap_colors = {"小": "DAEEF3", "中": P2COL, "大": "FCE4D6"}

    for i, m in enumerate(MODULES):
        r = i + 3
        (no, cn, en, domain, pri, mkt, comp, cur, tgt, gap,
         competitor, foundation, dur, team, pm, note) = m

        fill = priority_colors.get(pri, P3COL)
        gap_fill = gap_colors.get(gap, P3COL)

        # Row background
        for col in range(1, 17):
            ws.cell(row=r, column=col).fill = _fill(fill)

        data_cell(ws, r, 1,  no,         fill, align="center")
        data_cell(ws, r, 2,  cn,         fill, bold=(pri == "P1"))
        data_cell(ws, r, 3,  en,         fill)
        data_cell(ws, r, 4,  domain,     fill)

        # Priority badge
        pc = ws.cell(row=r, column=5, value=pri)
        pc.font = Font(name="Arial", bold=True,
                       color=WHITE if pri == "P1" else DNAV, size=10)
        pc.fill = _fill(MBLUE if pri == "P1" else (BBLUE if pri == "P2" else "9DC3E6"))
        pc.alignment = _align("center")
        pc.border = _border()

        data_cell(ws, r, 6,  "★"*mkt + "☆"*(5-mkt), fill, align="center")
        data_cell(ws, r, 7,  "★"*comp + "☆"*(5-comp), fill, align="center")
        data_cell(ws, r, 8,  cur,        fill, align="center")
        data_cell(ws, r, 9,  tgt,        fill, align="center")

        gc = ws.cell(row=r, column=10, value=gap)
        gc.font = Font(name="Arial", bold=True, size=10,
                       color=RED if gap == "大" else DNAV)
        gc.fill = _fill(gap_fill)
        gc.alignment = _align("center")
        gc.border = _border()

        data_cell(ws, r, 11, competitor, fill)
        data_cell(ws, r, 12, foundation, fill)
        data_cell(ws, r, 13, dur,        fill, align="center")
        data_cell(ws, r, 14, team,       fill, align="center")

        pmc = ws.cell(row=r, column=15, value=pm)
        pmc.font = Font(name="Arial", color="0000FF", size=10)
        pmc.fill = _fill(fill)
        pmc.alignment = _align("center")
        pmc.border = _border()

        data_cell(ws, r, 16, note, fill)
        ws.row_dimensions[r].height = 28

    # Totals row
    tr = len(MODULES) + 3
    ws.merge_cells(f"A{tr}:L{tr}")
    tc = ws.cell(row=tr, column=1, value="合计 / Totals")
    tc.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("right")
    tc.border = _border("4472C4")

    for col, formula in [
        (13, f"=SUM(M3:M{tr-1})"),
        (14, f"=MAX(N3:N{tr-1})"),
        (15, f"=SUM(O3:O{tr-1})"),
    ]:
        c = ws.cell(row=tr, column=col, value=formula)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        c.border = _border("4472C4")

    note_c = ws.cell(row=tr, column=16,
                     value="P1: 8模块(6项快赢+2项核心)  |  P2: 6模块  |  P3: 2模块")
    note_c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
    note_c.fill = _fill(DNAV)
    note_c.alignment = _align()
    note_c.border = _border("4472C4")

    ws.print_area = f"A1:P{tr}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


def make_s2(wb):
    ws = wb.create_sheet("关键功能清单")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:H1")
    t = ws.cell(row=1, column=1, value="FICC各模块关键功能清单  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    headers2 = [
        ("系统/模块", 18), ("优先级", 7), ("功能编号", 9), ("功能名称", 20),
        ("功能描述", 45), ("验收标准关键词", 20), ("复杂度", 8), ("预估工期(人天)", 12),
    ]
    for col, (h, w) in enumerate(headers2, 1):
        header_style(ws, 2, col, h, width=w)
    ws.row_dimensions[2].height = 26

    mod_lookup = {m[0]: m for m in MODULES}
    priority_colors = {"P1": P1COL, "P2": P2COL, "P3": P3COL}
    accept_kw = {
        "极高": "p99延迟/误差率/精度验证/压测10k TPS",
        "高":   "功能完整/边界测试/集成冒烟/性能基线",
        "中":   "功能测试通过/正确率100%",
        "低":   "功能测试通过",
    }

    cur_mod = None
    row = 3
    for (mod_no, func_no, func_cn, func_desc, complexity, dev_days) in FUNCTIONS:
        m = mod_lookup[mod_no]
        pri = m[4]
        fill = priority_colors.get(pri, P3COL)

        if cur_mod != mod_no:
            cur_mod = mod_no
            ws.merge_cells(f"A{row}:H{row}")
            label = f"{'【P1】' if pri=='P1' else '【'+pri+'】'} {m[1]}  ({m[2]})"
            gc = ws.cell(row=row, column=1, value=label)
            gc.font = Font(name="Arial", bold=True, size=10,
                           color=WHITE if pri == "P1" else DNAV)
            gc.fill = _fill(MBLUE if pri == "P1" else DNAV if pri == "P2" else "9DC3E6")
            gc.alignment = _align()
            gc.border = _border("4472C4")
            ws.row_dimensions[row].height = 20
            row += 1

        data_cell(ws, row, 1, m[1],      fill)
        data_cell(ws, row, 2, pri,        fill, align="center")
        data_cell(ws, row, 3, f"F{mod_no:02d}-{func_no:02d}", fill, align="center")
        data_cell(ws, row, 4, func_cn,   fill, bold=True)
        data_cell(ws, row, 5, func_desc,  fill)
        data_cell(ws, row, 6, accept_kw.get(complexity, ""), fill)

        cc = ws.cell(row=row, column=7, value=complexity)
        cc.font = Font(name="Arial", bold=True, size=9,
                       color=WHITE if complexity == "极高" else RED if complexity == "高" else DNAV)
        cc.fill = _fill(MBLUE if complexity == "极高" else
                        "D6E4F7" if complexity == "中" else CANVAS)
        cc.alignment = _align("center")
        cc.border = _border()

        dc = ws.cell(row=row, column=8, value=dev_days)
        dc.font = Font(name="Arial", color="0000FF", size=10)
        dc.fill = _fill(fill)
        dc.alignment = _align("center")
        dc.border = _border()

        ws.row_dimensions[row].height = 28
        row += 1

    # Totals footer
    ws.merge_cells(f"A{row}:G{row}")
    tf = ws.cell(row=row, column=1, value="总计功能数 / Total Function Points")
    tf.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tf.fill = _fill(DNAV)
    tf.alignment = _align("right")
    tf.border = _border("4472C4")

    tc = ws.cell(row=row, column=8, value=f"=SUM(H3:H{row-1})")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("center")
    tc.border = _border("4472C4")


def make_s3(wb):
    ws = wb.create_sheet("资源规划")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    t = ws.cell(row=1, column=1, value="FICC平台建设资源规划（三期）  —  华锐技术 / 2026")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    # Phase merged headers (row 2)
    ws.merge_cells("A2:A3")
    ws.merge_cells("B2:B3")
    ws.merge_cells("C2:E2")
    ws.merge_cells("F2:H2")
    ws.merge_cells("I2:K2")
    ws.merge_cells("L2:L3")

    for cell_ref, label, width in [
        ("A2", "岗位", 22), ("B2", "职责说明", 30),
        ("C2", "Phase 1（M1-M10）\nP1模块 — 核心平台建设", None),
        ("F2", "Phase 2（M11-M20）\nP2模块 — 能力扩展", None),
        ("I2", "Phase 3（M21-M27）\nP3模块 — 战略深水区", None),
        ("L2", "总招募\n峰值人力", 12),
    ]:
        c = ws[cell_ref]
        c.value = label
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        if width:
            ws.column_dimensions[cell_ref[0]].width = width

    # Sub-headers row 3
    for start_col in [3, 6, 9]:
        for j, sh in enumerate(["招募(人)", "工期(月)", "人月"]):
            col = start_col + j
            c = ws.cell(row=3, column=col, value=sh)
            c.font = Font(name="Arial", bold=True, color=WHITE, size=9)
            c.fill = _fill(MBLUE)
            c.alignment = _align("center")
            c.border = _border()
            ws.column_dimensions[get_column_letter(col)].width = 9

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 20

    p1_dur, p2_dur, p3_dur = 10, 10, 7

    for i, (role_cn, role_en, p1_hc, p2_hc, p3_hc, desc) in enumerate(ROLES):
        r = i + 4
        fill = P1COL if p1_hc >= 10 else P2COL

        data_cell(ws, r, 1,  f"{role_cn}\n{role_en}", fill, bold=True)
        data_cell(ws, r, 2,  desc, fill)

        for start_col, hc, dur in [(3, p1_hc, p1_dur), (6, p2_hc, p2_dur), (9, p3_hc, p3_dur)]:
            hc_c = ws.cell(row=r, column=start_col, value=hc)
            hc_c.font = Font(name="Arial", color="0000FF", size=10)
            hc_c.fill = _fill(fill)
            hc_c.alignment = _align("center")
            hc_c.border = _border()

            dur_c = ws.cell(row=r, column=start_col+1, value=dur)
            dur_c.font = Font(name="Arial", size=10)
            dur_c.fill = _fill(fill)
            dur_c.alignment = _align("center")
            dur_c.border = _border()

            hc_letter  = get_column_letter(start_col)
            dur_letter = get_column_letter(start_col + 1)
            pm_c = ws.cell(row=r, column=start_col+2,
                           value=f"={hc_letter}{r}*{dur_letter}{r}")
            pm_c.font = Font(name="Arial", size=10)
            pm_c.fill = _fill(fill)
            pm_c.alignment = _align("center")
            pm_c.border = _border()

        max_c = ws.cell(row=r, column=12, value=f"=MAX(C{r},F{r},I{r})")
        max_c.font = Font(name="Arial", bold=True, size=10)
        max_c.fill = _fill(fill)
        max_c.alignment = _align("center")
        max_c.border = _border()

        ws.row_dimensions[r].height = 32

    # Totals
    tr = len(ROLES) + 4
    ws.merge_cells(f"A{tr}:B{tr}")
    tc = ws.cell(row=tr, column=1, value="合计 Totals")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("right")
    tc.border = _border("4472C4")

    for col in [3, 6, 9, 12]:
        c = ws.cell(row=tr, column=col,
                    value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr-1})")
        c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        c.fill = _fill(DNAV)
        c.alignment = _align("center")
        c.border = _border("4472C4")

    for col in [4, 5, 7, 8, 10, 11]:
        c = ws.cell(row=tr, column=col, value="")
        c.fill = _fill(DNAV)
        c.border = _border("4472C4")

    # Notes
    nr = tr + 2
    ws.merge_cells(f"A{nr}:L{nr}")
    nc = ws.cell(row=nr, column=1, value=(
        "规划假设说明：① 各阶段人力可累计复用（P2团队延续P1人员，适当扩张）  "
        "② 量化研究员稀缺，建议优先招募，外部顾问补充  "
        "③ 信创合规专家建议从监管/银行方引进  "
        "④ 峰值人力约40人（Phase 2），建议2026Q3启动Phase 1招募"
    ))
    nc.font = Font(name="Arial", size=9, italic=True, color=DNAV)
    nc.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[nr].height = 40


def make_s4(wb):
    ws = wb.create_sheet("P1优先启动计划")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value="P1重点模块优先启动计划  —  华锐技术 FICC / 2026-2027")
    t.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")

    headers = [
        ("模块", 18), ("启动月份", 9), ("验收里程碑", 22), ("首个客户场景", 22),
        ("核心技术风险", 22), ("风险缓解策略", 22), ("负责团队", 14),
        ("工期(月)", 8), ("团队(人)", 8), ("快赢商业价值", 22),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 2, col, h, width=w)
    ws.row_dimensions[2].height = 26

    p1_data = [
        (1,  "异常交易风控",     "M1",  "上线：10支债券实时监控告警，0漏报",   "某券商债券异常订单旁路检测",
             "CEP引擎低延迟稳定性",     "规则引擎轻量先行，ML模型后置",     "风控平台组", 6, 8,
             "监管驱动，客户愿付费，快速复制推广"),
        (2,  "事前风控/合规",    "M1",  "上线：200+合规规则覆盖，0漏拦",      "某券商自营债券投资合规校验",
             "规则库完整性",           "与合规团队联合建规则",             "风控平台组", 5, 6,
             "与异常风控组合销售，提升ARPU"),
        (3,  "现券交易系统",     "M2",  "上线：银行间+交易所双市场报盘",       "替换某券商现有债券交易终端",
             "CFETS接口稳定性",        "早期接入测试环境联调",             "交易系统组", 6, 8,
             "替换恒生/金证，直接竞争替代"),
        (4,  "回购管理",         "M3",  "上线：正回购全周期，质押替换",        "某券商质押式回购融资管理",
             "中登/上清所接口联调",    "建立完整沙箱测试套件",              "交易系统组", 5, 6,
             "与现券交易捆绑，提升粘性"),
        (5,  "IBOR投资账薄",     "M2",  "上线：实时持仓净值，T0对账零差异",   "资管公司债券持仓实时账薄",
             "多资产估值覆盖率",        "分批上线（先债券，后衍生品）",     "平台基础组", 8, 10,
             "平台锁定效应，长期替换衡泰IBOR"),
        (6,  "组合管理+绩效归因","M4",  "上线：Brinson归因，周报自动生成",    "资管公司固收组合绩效分析",
             "Brinson模型精度验证",    "对标Bloomberg绩效分析基准",         "投资管理组", 9, 10,
             "高端客户差异化，替代SimCorp/衡泰"),
        (7,  "策略投资/量化",    "M5",  "上线：回测引擎，3个策略实盘验证",    "自营量化部门策略回测平台",
             "Tick回测数据质量",        "先与AMD行情平台深度整合",          "量化研究组", 10, 12,
             "高壁垒护城河，未来SaaS化"),
        (8,  "FICC风险计量",     "M3",  "上线：VaR/压测，监管格式报告",       "某券商FICC组合风险日报",
             "模型精度vs衡泰基准",      "首期仅VaR/ES，CVA后置",            "风控平台组", 10, 10,
             "直接替代衡泰，高溢价定价"),
    ]

    for i, row_data in enumerate(p1_data):
        r = i + 3
        (seq, mod, start, milestone, first_case,
         risk, mitigation, team, dur, hc, value) = row_data

        fill = P1COL
        for col in range(1, 11):
            ws.cell(row=r, column=col).fill = _fill(fill)

        data_cell(ws, r, 1,  f"{seq}. {mod}", fill, bold=True)

        mc = ws.cell(row=r, column=2, value=start)
        mc.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        mc.fill = _fill(MBLUE)
        mc.alignment = _align("center")
        mc.border = _border()

        data_cell(ws, r, 3,  milestone,   fill)
        data_cell(ws, r, 4,  first_case,  fill)

        rc = ws.cell(row=r, column=5, value=risk)
        rc.font = Font(name="Arial", size=10, color=RED)
        rc.fill = _fill(fill)
        rc.alignment = _align(wrap=True)
        rc.border = _border()

        data_cell(ws, r, 6,  mitigation, fill)
        data_cell(ws, r, 7,  team,       fill, align="center")

        for col, val in [(8, dur), (9, hc)]:
            vc = ws.cell(row=r, column=col, value=val)
            vc.font = Font(name="Arial", color="0000FF", size=10)
            vc.fill = _fill(fill)
            vc.alignment = _align("center")
            vc.border = _border()

        data_cell(ws, r, 10, value, fill, bold=True)
        ws.row_dimensions[r].height = 44

    # Strategy summary box
    ir = len(p1_data) + 3 + 1
    ws.merge_cells(f"A{ir}:J{ir}")
    ic = ws.cell(row=ir, column=1, value=(
        "P1阶段战略重心：① 快赢先行（异常风控+事前合规M1并行，6个月见成果）  "
        "② 主干平行（IBOR+现券M2启动，建立账薄主干）  "
        "③ 差异化（量化策略+绩效归因M4-M5，构建高壁垒能力）  "
        "④ 替代衡泰（FICC风险计量M3启动，以信创+精度优势直攻垄断盲区）"
    ))
    ic.font = Font(name="Arial", size=9, bold=True, color=DNAV)
    ic.fill = _fill(CFILL)
    ic.alignment = Alignment(wrap_text=True, vertical="center")
    ic.border = _border(DNAV)
    ws.row_dimensions[ir].height = 52


def make_detail_portfolio(wb):
    """Dedicated detail sheet for 组合管理+绩效归因 — sourced from Slide 44 of
    基于国内外FICC平台发展经验规划FICC平台建设路径0417.pptx"""

    ws = wb.create_sheet("详情_组合管理")
    ws.sheet_view.showGridLines = False

    # ── Title ──────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1,
                value="系统详情：组合管理+绩效归因  |  Portfolio Management & Performance Attribution")
    t.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    t.fill = _fill(NAVY)
    t.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    # ── Core Objective (from Slide 44) ─────────────────────────────────────────
    ws.merge_cells("A2:J2")
    obj = ws.cell(row=2, column=1, value=(
        "核心目标（Slide 44）：为FICC投资组合管理提供全面解决方案，"
        "通过分析投资组合头寸、风险计量、绩效归因，"
        "为投资经理在资产管理、组合构建、组合再平衡、指数投资等方面赋能。"
    ))
    obj.font = Font(name="Arial", size=10, italic=True, color=DNAV)
    obj.fill = _fill(CFILL)
    obj.alignment = _align(wrap=True)
    obj.border = _border(DNAV)
    ws.row_dimensions[2].height = 36

    # ── Architecture Block (data inputs → quant base → analysis platform) ─────
    ws.merge_cells("A3:J3")
    arch = ws.cell(row=3, column=1, value=(
        "架构层次：  "
        "【数据层】交易数据 / 头寸数据 / 行情数据 / 估值数据  →  "
        "【量化底座】定价计算 · 利率曲线 · 因子模型  →  "
        "【组合分析平台】6大功能域（见下表）"
    ))
    arch.font = Font(name="Arial", size=10, bold=True, color=WHITE)
    arch.fill = _fill(DNAV)
    arch.alignment = _align(wrap=True)
    arch.border = _border("4472C4")
    ws.row_dimensions[3].height = 28

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = [
        ("功能域", 16), ("编号", 8), ("功能名称", 22), ("功能描述（来源Slide44）", 48),
        ("验收标准", 26), ("集成依赖", 20), ("复杂度", 8),
        ("预估(人天)", 10), ("优先级", 7), ("备注", 18),
    ]
    for col, (h, w) in enumerate(headers, 1):
        header_style(ws, 4, col, h, width=w)
    ws.row_dimensions[4].height = 28

    # ── Function data: (domain, domain_fill, func_no, name, desc, acceptance,
    #                    dependency, complexity, days, priority, note)
    funcs = [
        # A. 组合分析
        ("A. 组合分析", MBLUE,
         "F07-01", "债券横截面与时序分析",
         "跨品种横截面信号挖掘（利差/久期偏差），历史时序趋势与均值回归分析，支持日/周/月时间窗口",
         "横截面信号与实际收益相关性>0.3；时序模型回测夏普>0.5",
         "AMD行情平台 / 估值数据",
         "高", 35, "P1", "Slide44首项分析能力"),
        ("A. 组合分析", MBLUE,
         "F07-02", "相对价值分析",
         "利差/Z-spread/OAS相对价值计算，历史分位数定位，跨品种套利信号识别，个券评分",
         "OAS计算误差<1bps；分位数计算覆盖5年历史",
         "量化定价引擎 / AMD行情",
         "高", 30, "P1", "投资经理最常用分析工具"),
        ("A. 组合分析", MBLUE,
         "F07-03", "组合相关度分析",
         "资产间收益相关矩阵计算（滚动窗口60/120/250日），组合集中度风险识别，分散化系数监控",
         "相关矩阵计算耗时<30s（500券组合）",
         "IBOR头寸 / AMD历史行情",
         "中", 25, "P1", ""),

        # B. 风险计量
        ("B. 风险计量", DNAV,
         "F07-04", "情景分析",
         "利率平移/扭转/蝶式、信用利差扩大、流动性冲击三类情景；自定义情景编辑器；多情景P&L并行对比",
         "情景P&L计算与手工核算误差<0.5%；支持≥20个情景并行",
         "FICC风险计量平台 / 量化定价引擎",
         "高", 35, "P1", "复用风险计量模块情景库"),
        ("B. 风险计量", DNAV,
         "F07-05", "敏感度分析",
         "久期/DV01/凸度/利差/Vega敏感度多因子并行计算，组合层面聚合，对冲比率建议输出",
         "DV01与Bloomberg基准误差<0.5bps",
         "量化定价引擎Greeks接口",
         "高", 25, "P1", ""),
        ("B. 风险计量", DNAV,
         "F07-06", "极端事件分析",
         "内置924行情/2015股灾/2020疫情等8+极端情景，组合尾部损失分布，ES(99%)计算",
         "历史情景复现误差<1%；与风险计量平台情景库共享",
         "FICC风险计量平台",
         "高", 30, "P1", "与风险计量平台共享情景库"),

        # C. 绩效归因
        ("C. 绩效归因", BBLUE,
         "F07-07", "指数比较分析",
         "对标中债综合/信用/政策行/定制基准指数；跟踪误差（TE）日频实时监控；超额Alpha三因子分解",
         "TE计算误差<1bps；支持≥10个基准指数同时跟踪",
         "AMD中债估值 / IBOR持仓",
         "中", 25, "P1", "中债指数数据接入为前提"),
        ("C. 绩效归因", BBLUE,
         "F07-08", "绩效归因（Brinson模型）",
         "BHB三效应：资产配置/个券选择/交互效应；日/周/月/季多时间维度；报告PDF自动生成",
         "Brinson归因结果与Bloomberg AIM误差<0.1%；报告T+1 9:00前自动生成",
         "IBOR持仓 / AMD行情 / 指数数据",
         "高", 40, "P1", "核心差异化功能，对标SimCorp/Aladdin"),

        # D. 组合优化与再平衡
        ("D. 组合优化与再平衡", "3B7DD8",
         "F07-09", "虚拟组合分析",
         "组合调整前后净值/风险/收益效果模拟（纸面交易）；市场冲击成本预估；多方案结果对比展示",
         "虚拟组合模拟结果与真实执行偏差<0.3%",
         "IBOR / 现券交易系统（冲击模型）",
         "高", 20, "P1", ""),
        ("D. 组合优化与再平衡", "3B7DD8",
         "F07-10", "再平衡方案试算与比较",
         "久期/集中度/流动性/合规约束联合优化（二次规划QP），多方案并行试算，最优路径推荐",
         "QP求解100券组合<5s；约束满足率100%",
         "事前风控/合规 / IBOR / 现券系统",
         "高", 35, "P1", "QP求解器选型：CVXPY/Gurobi"),
        ("D. 组合优化与再平衡", "3B7DD8",
         "F07-11", "资产配置策略分析",
         "宏观因子（利率周期/信用周期/流动性）驱动大类资产配置模型；动态权重优化建议；历史回测验证",
         "回测期≥5年；模型收益预测方向胜率>55%",
         "AMD宏观数据 / 因子模型",
         "高", 40, "P2", "Phase 1后期或Phase 2"),
        ("D. 组合优化与再平衡", "3B7DD8",
         "F07-12", "固收+策略分析",
         "债券底仓+衍生品增强策略P&L归因，风险分解（利率/信用/期权贡献），固收+产品净值模拟",
         "固收+净值模拟与实际偏差<0.5%",
         "IBOR / 利率衍生品系统（Phase 2）",
         "高", 30, "P2", "依赖利率衍生品模块"),
        ("D. 组合优化与再平衡", "3B7DD8",
         "F07-13", "量化策略组合分析",
         "多策略组合叠加容量约束，策略相关性管理，Sharpe/最大回撤/Calmar实时监控，策略权重优化",
         "策略容量计算误差<5%；组合Sharpe≥1.2（回测）",
         "策略投资/量化系统",
         "高", 35, "P2", "与策略投资模块深度整合"),

        # E. 流动性
        ("E. 流动性管理", "9DC3E6",
         "F07-14", "流动性分析",
         "逐券换手率/市场深度评估（日均成交量法），组合变现能力分析，流动性压力下变现损失（LVaR）估算",
         "LVaR计算与市场基准误差<5%；日均成交量数据覆盖≥95%持仓",
         "AMD行情（成交量） / IBOR持仓",
         "中", 20, "P1", "监管流动性要求驱动"),

        # F. 量化底座
        ("F. 量化底座集成", CANVAS,
         "F07-15", "定价计算集成",
         "调用量化定价引擎实时估值与Greeks；支持含权债/可转债/IRS；结果缓存（Redis）与Kafka推送",
         "定价API p99<200ms；日终全量估值<30min（10万券）",
         "量化定价引擎（P2模块）",
         "高", 20, "P2", "Phase 1用AMD估值代替，Phase 2升级"),
        ("F. 量化底座集成", CANVAS,
         "F07-16", "利率曲线集成",
         "多曲线实时消费（国债/政策行/AAA信用/OIS）；Nelson-Siegel参数提取（水平/斜率/曲率）；敏感度映射",
         "曲线更新延迟<1s；NS参数拟合R²>0.999",
         "AMD行情平台 / 量化定价引擎",
         "中", 15, "P1", ""),
        ("F. 量化底座集成", CANVAS,
         "F07-17", "因子模型",
         "多因子风险模型（利率/信用/流动性/行业因子）；Alpha因子库（≥20个因子）；因子暴露实时计算与归因",
         "因子模型解释方差>80%；Alpha IC均值>0.05",
         "AMD历史行情 / IBOR持仓",
         "高", 40, "P1", "复用策略投资模块因子库"),
    ]

    domain_fill_map = {}
    row = 5
    cur_domain = None

    for item in funcs:
        (domain, dom_fill, fno, fname, fdesc, accept, dep,
         complexity, days, priority, note) = item

        # Domain group header
        if cur_domain != domain:
            cur_domain = domain
            ws.merge_cells(f"A{row}:J{row}")
            dh = ws.cell(row=row, column=1, value=domain)
            is_dark = dom_fill not in (CANVAS, "9DC3E6")
            dh.font = Font(name="Arial", bold=True, size=10,
                           color=WHITE if is_dark else DNAV)
            dh.fill = _fill(dom_fill)
            dh.alignment = _align()
            dh.border = _border("4472C4")
            ws.row_dimensions[row].height = 18
            row += 1

        # Row fill: alternate light shades per domain
        row_fill = P1COL if priority == "P1" else P2COL

        data_cell(ws, row, 1, domain,     row_fill)
        data_cell(ws, row, 2, fno,        row_fill, align="center")
        data_cell(ws, row, 3, fname,      row_fill, bold=True)
        data_cell(ws, row, 4, fdesc,      row_fill)
        data_cell(ws, row, 5, accept,     row_fill)
        data_cell(ws, row, 6, dep,        row_fill)

        cc = ws.cell(row=row, column=7, value=complexity)
        cc.font = Font(name="Arial", bold=True, size=9,
                       color=WHITE if complexity == "极高" else RED if complexity == "高" else DNAV)
        cc.fill = _fill(MBLUE if complexity == "极高" else P2COL)
        cc.alignment = _align("center")
        cc.border = _border()

        dc = ws.cell(row=row, column=8, value=days)
        dc.font = Font(name="Arial", color="0000FF", size=10)
        dc.fill = _fill(row_fill)
        dc.alignment = _align("center")
        dc.border = _border()

        pc = ws.cell(row=row, column=9, value=priority)
        pc.font = Font(name="Arial", bold=True, size=10,
                       color=WHITE if priority == "P1" else DNAV)
        pc.fill = _fill(MBLUE if priority == "P1" else BBLUE)
        pc.alignment = _align("center")
        pc.border = _border()

        data_cell(ws, row, 10, note, row_fill)
        ws.row_dimensions[row].height = 48
        row += 1

    # Totals
    ws.merge_cells(f"A{row}:G{row}")
    tf = ws.cell(row=row, column=1,
                 value=f"合计 17项功能  |  P1: 12项  |  P2: 5项  |  来源：Slide 44 + 扩展")
    tf.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tf.fill = _fill(DNAV)
    tf.alignment = _align()
    tf.border = _border("4472C4")

    tc = ws.cell(row=row, column=8, value=f"=SUM(H5:H{row-1})")
    tc.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    tc.fill = _fill(DNAV)
    tc.alignment = _align("center")
    tc.border = _border("4472C4")

    for col in [9, 10]:
        c = ws.cell(row=row, column=col, value="")
        c.fill = _fill(DNAV)
        c.border = _border("4472C4")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1


def main():
    wb = Workbook()
    wb.remove(wb.active)

    make_s1(wb)
    make_s2(wb)
    make_s3(wb)
    make_s4(wb)
    make_detail_portfolio(wb)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
